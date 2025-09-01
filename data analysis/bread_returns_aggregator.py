# -*- coding: utf-8 -*-
"""
Bread Factory – GUI scanner (FAST; bounded A1:Y320; robust to merged headers)

What it does:
- Lets user choose an .xlsx file and a date range (DD.MM.YYYY)
- Scans only sheets whose names match DD.MM.YYYY and fall within the date range
- In each sheet, searches only A1:Y320 for any cell that equals "Вид Хлеба"
- For each found table:
    * Store names are on the row ABOVE "Вид Хлеба"
    * Subheaders row (same row as "Вид Хлеба") repeats:  П-ка | В-ат | Ост
    * Skips "К-во Хлеба" column (just to the right of "Вид Хлеба")
    * Ignores "Ост", ignores "К-во Хлеба"
    * Sums П-ка (brought) and В-ат (returned)
    * If the cell above "Вид Хлеба" has a driver name -> adds to driver totals
      Otherwise the table counts only for stores (not for drivers)
- Displays results in two tables (Stores / Drivers) with return rate %
- Exports CSV or Excel with one click

Notes:
- Designed to be SAFE and FAST: only A..Y, rows 1..320 are scanned.
- Handles merged cells (store names, driver name) by reading the top-left value.
"""

from __future__ import annotations
import re
import csv
import unicodedata
from dataclasses import dataclass
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


# ---------------- Configuration (tweak if needed) ----------------
MAX_ROWS = 320          # rows 1..320 only
MAX_COLS = 25           # A..Y => 25
ANCHOR_TEXT = "вид хлеба"
SUB_PKA = "п-ка"
SUB_VAT = "в-ат"
# ----------------------------------------------------------------


@dataclass
class StoreCols:
    name: str
    pka_col: int
    vat_col: int


# ------------- Core helpers (pure logic; no GUI) -----------------
def norm_text(v) -> str:
    """Normalize text for robust comparison (case/space/hyphen insensitive)."""
    if v is None:
        return ""
    s = str(v)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\u2011", "-").replace("\u2013", "-").replace("\u2014", "-")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def to_number(v) -> float:
    """Convert various cell contents to float; non-numeric -> 0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else 0.0


def is_date_sheet(name: str) -> Optional[datetime]:
    """Return datetime if sheet name is DD.MM.YYYY, else None."""
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", name or ""):
        try:
            return datetime.strptime(name, "%d.%m.%Y")
        except ValueError:
            return None
    return None


def cell_top_left_of_merge(ws: Worksheet, row: int, col: int) -> Tuple[int, int]:
    """If (row,col) is inside a merged range, return that range's top-left."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr.min_row, mr.min_col
    return row, col


def read_cell(ws: Worksheet, row: int, col: int):
    """Read a cell, resolving merged cells to their top-left value."""
    r0, c0 = cell_top_left_of_merge(ws, row, col)
    return ws.cell(r0, c0).value


def find_table_anchors(ws: Worksheet, last_row: int, last_col: int) -> List[Tuple[int, int]]:
    """Find all cells within A1..(last_row,last_col) that equal 'Вид Хлеба'."""
    anchors = []
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            if norm_text(read_cell(ws, r, c)) == ANCHOR_TEXT:
                anchors.append((r, c))
    return anchors


def detect_store_columns(ws: Worksheet, hdr_row: int, sub_row: int,
                         start_col: int, last_col: int) -> List[StoreCols]:
    """Detect repeating [П-ка, В-ат, (Ост)] blocks with store names above."""
    stores: List[StoreCols] = []
    c = start_col
    while c <= last_col - 1:
        pka = norm_text(read_cell(ws, sub_row, c))
        vat = norm_text(read_cell(ws, sub_row, c + 1))
        if pka == SUB_PKA and vat == SUB_VAT:
            name_raw = read_cell(ws, hdr_row, c)
            pretty = str(name_raw).strip() if name_raw else ""
            stores.append(StoreCols(pretty or "(без названия)", c, c + 1))
            c += 3  # skip 'Ост' column
        else:
            c += 1
    return stores


def data_row_end(ws: Worksheet, bread_col: int, start_row: int, last_row: int) -> int:
    """
    Determine last data row by scanning column 'bread_col'.
    Stop after two consecutive empty rows (safety) or end of window.
    """
    r = start_row
    blank = 0
    while r <= last_row:
        if norm_text(read_cell(ws, r, bread_col)) == "":
            blank += 1
            if blank >= 2:
                break
        else:
            blank = 0
        r += 1
    return r - blank - 1


def scan_workbook(path: Path, date_from: datetime, date_to: datetime,
                  max_rows: int = MAX_ROWS, max_cols: int = MAX_COLS,
                  progress_cb=None, log_cb=None):
    """
    Main scanning routine; returns (stores_totals, drivers_totals, meta_dict).
    Callbacks:
      - progress_cb(done, total) -> None
      - log_cb(str) -> None
    """
    stores_totals: Dict[str, Dict[str, float]] = defaultdict(lambda: {"brought": 0.0, "returned": 0.0})
    drivers_totals: Dict[str, Dict[str, float]] = defaultdict(lambda: {"brought": 0.0, "returned": 0.0})

    wb = load_workbook(path, data_only=True)
    sheets = [n for n in wb.sheetnames if (dt := is_date_sheet(n)) and date_from <= dt <= date_to]
    total = len(sheets)
    done = 0
    if log_cb:
        log_cb(f"Scanning {total} sheet(s) between {date_from:%d.%m.%Y} and {date_to:%d.%m.%Y}…")

    for name in sheets:
        ws: Worksheet = wb[name]
        # Boundaries
        last_row = min(ws.max_row or max_rows, max_rows)
        last_col = min(ws.max_column or max_cols, max_cols)

        anchors = find_table_anchors(ws, last_row, last_col)
        if log_cb:
            lim = f"A1:{get_column_letter(last_col)}{last_row}"
            log_cb(f"  [{name}] found {len(anchors)} table(s) in {lim}")

        for (anchor_row, anchor_col) in anchors:
            # Driver name is above the "Вид Хлеба" cell
            driver_raw = read_cell(ws, anchor_row - 1, anchor_col)
            driver_display = str(driver_raw).strip() if driver_raw else ""
            has_driver = bool(norm_text(driver_raw))

            # Rows/cols
            hdr_row = anchor_row - 1             # store names
            sub_row = anchor_row                 # П-ка/В-ат/Ост
            data_start = anchor_row + 1
            bread_col = anchor_col
            first_store_col = anchor_col + 2     # skip 'К-во Хлеба'

            stores = detect_store_columns(ws, hdr_row, sub_row, first_store_col, last_col)
            if not stores:
                continue

            data_end = data_row_end(ws, bread_col, data_start, last_row)
            if data_end < data_start:
                continue

            # Sum all rows
            for r in range(data_start, data_end + 1):
                for st in stores:
                    b = to_number(read_cell(ws, r, st.pka_col))
                    rv = to_number(read_cell(ws, r, st.vat_col))

                    # Stores (always)
                    stores_totals[st.name]["brought"] += b
                    stores_totals[st.name]["returned"] += rv

                    # Drivers (only if exists)
                    if has_driver:
                        drivers_totals[driver_display]["brought"] += b
                        drivers_totals[driver_display]["returned"] += rv

        done += 1
        if progress_cb:
            progress_cb(done, total)

    meta = {
        "sheets_scanned": total,
        "path": str(path),
        "from": date_from.strftime("%d.%m.%Y"),
        "to": date_to.strftime("%d.%m.%Y"),
        "max_rows": max_rows,
        "max_cols": max_cols,
    }
    return stores_totals, drivers_totals, meta


def rate(returned: float, brought: float) -> float:
    return (returned / brought * 100.0) if brought > 0 else 0.0


# ------------------------------ GUI ------------------------------
class BreadGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("Bread Factory Scanner")
        root.geometry("980x640")

        # Top controls
        pad = {"padx": 6, "pady": 6}
        frm = ttk.Frame(root)
        frm.pack(fill="x", **pad)

        self.path_var = tk.StringVar()
        self.from_var = tk.StringVar()
        self.to_var = tk.StringVar()
        self.rows_var = tk.IntVar(value=MAX_ROWS)
        self.cols_var = tk.IntVar(value=MAX_COLS)

        ttk.Label(frm, text="Workbook:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.path_var, width=60).grid(row=0, column=1, sticky="we")
        ttk.Button(frm, text="Browse…", command=self.choose_file).grid(row=0, column=2, sticky="w")

        ttk.Label(frm, text="Date from (DD.MM.YYYY):").grid(row=1, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.from_var, width=16).grid(row=1, column=1, sticky="w")
        ttk.Label(frm, text="to:").grid(row=1, column=1, sticky="e", padx=(0, 170))
        ttk.Entry(frm, textvariable=self.to_var, width=16).grid(row=1, column=2, sticky="w")

        ttk.Label(frm, text="Scan rows (max):").grid(row=2, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.rows_var, width=8).grid(row=2, column=1, sticky="w")
        ttk.Label(frm, text="Scan cols (max):").grid(row=2, column=1, sticky="e", padx=(0, 170))
        ttk.Entry(frm, textvariable=self.cols_var, width=8).grid(row=2, column=2, sticky="w")

        frm.grid_columnconfigure(1, weight=1)

        # Buttons
        btns = ttk.Frame(root)
        btns.pack(fill="x", **pad)
        self.scan_btn = ttk.Button(btns, text="Scan", command=self.start_scan)
        self.scan_btn.pack(side="left")
        ttk.Button(btns, text="Export CSV", command=self.export_csv).pack(side="left", padx=6)
        ttk.Button(btns, text="Export Excel", command=self.export_excel).pack(side="left")

        # Progress + log
        prog = ttk.Frame(root)
        prog.pack(fill="x", **pad)
        self.pb = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pb.pack(fill="x", expand=True, side="left")
        self.status = ttk.Label(prog, text="Idle")
        self.status.pack(side="left", padx=8)

        logf = ttk.Frame(root)
        logf.pack(fill="both", expand=True, **pad)

        # Tabs with results
        self.nb = ttk.Notebook(logf)
        self.nb.pack(fill="both", expand=True, side="left")

        self.tree_stores = self._make_tree(self.nb, ("Store", "Brought (П-ка)", "Returned (В-ат)", "Return %"))
        self.tree_drivers = self._make_tree(self.nb, ("Driver", "Brought (П-ка)", "Returned (В-ат)", "Return %"))
        self.nb.add(self.tree_stores, text="Stores")
        self.nb.add(self.tree_drivers, text="Drivers")

        # Log box
        self.log = tk.Text(logf, width=36)
        self.log.pack(fill="y", side="right")
        self._log("Ready.")

        # Data
        self.stores_totals = {}
        self.drivers_totals = {}
        self.last_meta = {}

        # Threading state
        self._worker_id = None

    def _make_tree(self, parent, headings):
        tree = ttk.Treeview(parent, columns=list(range(1, len(headings) + 1)), show="headings")
        for i, h in enumerate(headings, start=1):
            tree.heading(i, text=h)
            w = 240 if i == 1 else 140
            tree.column(i, width=w, anchor="e" if i > 1 else "w")
        vsb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        # Place tree in its own frame so the scrollbar shows correctly
        wrapper = ttk.Frame(parent)
        tree.pack(in_=wrapper, side="left", fill="both", expand=True)
        vsb.pack(in_=wrapper, side="right", fill="y")
        return wrapper

    def _get_tree_widget(self, wrapper):
        # helper to fetch the Treeview from the wrapper frame
        for ch in wrapper.winfo_children():
            if isinstance(ch, ttk.Treeview):
                return ch
        return None

    # ---------- UI handlers ----------
    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Choose Excel workbook",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.path_var.set(path)

    def start_scan(self):
        path = self.path_var.get().strip()
        if not path:
            messagebox.showerror("Error", "Please choose a workbook file.")
            return
        try:
            date_from = datetime.strptime(self.from_var.get().strip(), "%d.%m.%Y")
            date_to = datetime.strptime(self.to_var.get().strip(), "%d.%m.%Y")
        except Exception:
            messagebox.showerror("Error", "Please enter valid dates as DD.MM.YYYY.")
            return
        if date_to < date_from:
            messagebox.showerror("Error", "End date must be >= start date.")
            return

        max_rows = max(1, int(self.rows_var.get() or MAX_ROWS))
        max_cols = max(1, int(self.cols_var.get() or MAX_COLS))

        # Disable button, reset progress
        self.scan_btn.configure(state="disabled")
        self.pb["value"] = 0
        self.status.configure(text="Scanning…")
        self._log(f"Starting scan: {path}  |  {date_from:%d.%m.%Y} → {date_to:%d.%m.%Y}  |  Window A1:{get_column_letter(max_cols)}{max_rows}")

        # Run in an "after" loop (keeps UI responsive without threads)
        tasks = {"step": 0, "result": None, "error": None}

        def work():
            try:
                stores, drivers, meta = scan_workbook(
                    Path(path), date_from, date_to,
                    max_rows=max_rows, max_cols=max_cols,
                    progress_cb=self._progress_cb, log_cb=self._log
                )
                tasks["result"] = (stores, drivers, meta)
            except Exception as e:
                tasks["error"] = e
            finally:
                self.root.after(0, finished)

        def finished():
            self.scan_btn.configure(state="normal")
            if tasks["error"] is not None:
                self._log(f"ERROR: {tasks['error']}")
                messagebox.showerror("Scan failed", str(tasks["error"]))
                self.status.configure(text="Error")
                return

            (stores, drivers, meta) = tasks["result"]
            self.stores_totals, self.drivers_totals, self.last_meta = stores, drivers, meta
            self._populate_results()
            self.status.configure(text="Done")
            self._log("Scan complete.")

        # Use a very short "thread" via after() to not block the UI
        # (Tkinter is single-thread friendly with after callbacks)
        self.root.after(10, work)

    def _progress_cb(self, done, total):
        self.pb["maximum"] = max(1, total)
        self.pb["value"] = done

    def _log(self, msg: str):
        self.log.insert("end", msg + "\n")
        self.log.see("end")

    def _populate_results(self):
        # Clear
        store_tree = self._get_tree_widget(self.tree_stores)
        driver_tree = self._get_tree_widget(self.tree_drivers)
        for t in (store_tree, driver_tree):
            for i in t.get_children():
                t.delete(i)

        # Fill stores
        for s in sorted(self.stores_totals.keys()):
            b = self.stores_totals[s]["brought"]
            r = self.stores_totals[s]["returned"]
            rate_pct = rate(r, b)
            store_tree.insert("", "end", values=(s, self._fmt_num(b), self._fmt_num(r), f"{rate_pct:.2f}"))

        # Fill drivers
        for d in sorted(self.drivers_totals.keys()):
            b = self.drivers_totals[d]["brought"]
            r = self.drivers_totals[d]["returned"]
            rate_pct = rate(r, b)
            driver_tree.insert("", "end", values=(d, self._fmt_num(b), self._fmt_num(r), f"{rate_pct:.2f}"))

    @staticmethod
    def _fmt_num(x: float):
        return str(int(x)) if float(x).is_integer() else f"{x:.3f}"

    # ---------- Export ----------
    def export_csv(self):
        if not self.stores_totals and not self.drivers_totals:
            messagebox.showinfo("Nothing to export", "Run a scan first.")
            return
        base = filedialog.askdirectory(title="Choose folder to save CSV files")
        if not base:
            return
        base = Path(base)
        meta = self.last_meta or {}
        suffix = f"{meta.get('from','')}_{meta.get('to','')}"
        stores_path = base / f"stores_summary_{suffix}.csv"
        drivers_path = base / f"drivers_summary_{suffix}.csv"

        with stores_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Store", "Brought (П-ка)", "Returned (В-ат)", "Return rate %"])
            for s in sorted(self.stores_totals.keys()):
                b = self.stores_totals[s]["brought"]
                r = self.stores_totals[s]["returned"]
                w.writerow([s, self._fmt_num(b), self._fmt_num(r), f"{rate(r, b):.2f}"])

        with drivers_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Driver", "Brought (П-ка)", "Returned (В-ат)", "Return rate %"])
            for d in sorted(self.drivers_totals.keys()):
                b = self.drivers_totals[d]["brought"]
                r = self.drivers_totals[d]["returned"]
                w.writerow([d, self._fmt_num(b), self._fmt_num(r), f"{rate(r, b):* .2f}".replace("* ", "")])

        messagebox.showinfo("CSV exported", f"Saved:\n{stores_path}\n{drivers_path}")
        self._log(f"CSV saved: {stores_path}\nCSV saved: {drivers_path}")

    def export_excel(self):
        if not self.stores_totals and not self.drivers_totals:
            messagebox.showinfo("Nothing to export", "Run a scan first.")
            return
        fpath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Save results workbook as…",
            initialfile="bread_summary.xlsx",
        )
        if not fpath:
            return
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Stores"
        ws2 = wb.create_sheet("Drivers")

        # Headers
        ws1.append(["Store", "Brought (П-ка)", "Returned (В-ат)", "Return rate %"])
        ws2.append(["Driver", "Brought (П-ка)", "Returned (В-ат)", "Return rate %"])

        # Data
        for s in sorted(self.stores_totals.keys()):
            b = self.stores_totals[s]["brought"]
            r = self.stores_totals[s]["returned"]
            ws1.append([s, float(b), float(r), round(rate(r, b), 2)])

        for d in sorted(self.drivers_totals.keys()):
            b = self.drivers_totals[d]["brought"]
            r = self.drivers_totals[d]["returned"]
            ws2.append([d, float(b), float(r), round(rate(r, b), 2)])

        # Cosmetic
        for ws in (ws1, ws2):
            ws.freeze_panes = "A2"
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 22

        try:
            wb.save(fpath)
            messagebox.showinfo("Excel exported", f"Saved: {fpath}")
            self._log(f"Excel saved: {fpath}")
        except Exception as e:
            messagebox.showerror("Save failed", str(e))
            self._log(f"ERROR saving Excel: {e}")


# ------------------------------ main -----------------------------
def main():
    root = tk.Tk()
    # Native-like theme if available
    try:
        root.call("source", "azure.tcl")  # harmless if missing
        ttk.Style().theme_use("azure")
    except Exception:
        pass
    BreadGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
